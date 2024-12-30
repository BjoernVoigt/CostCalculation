import sys
import os
import math
import re
import pandas as pd
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QPushButton, QLabel, QFileDialog, QWidget, QSpacerItem, QSizePolicy,
    QHBoxLayout, QGridLayout, QGroupBox, QComboBox, QTabWidget, QLineEdit, QListWidget, QListWidgetItem, QRadioButton, QButtonGroup,
    QGraphicsOpacityEffect, QMessageBox, QFrame,
)
from PyQt5.QtGui import QIcon, QPixmap, QFontMetrics, QFont
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtCore import pyqtSignal, QTimer, QRect
from openpyxl.styles import Alignment
import FreeCAD
import Part
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import subprocess
import vtk
from vtkmodules.vtkRenderingCore import vtkRenderer, vtkActor, vtkPolyDataMapper
from vtkmodules.vtkIOPLY import vtkPLYReader
from vtkmodules.qt.QVTKRenderWindowInteractor import QVTKRenderWindowInteractor

# Specify if on desktop or laptop
desktop = True


########## Paths ##########
if desktop == True:
    colourprice = r"C:\Users\Bjørn\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\ColourPrice.xlsx"
    Logo = r"C:\Users\Bjørn\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\logo_3.png"
    res_excel = r"C:\Users\Bjørn\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\results3.xlsx"
    stp_folder = r"C:\Cost_Calculation\StepFiles\Case1 - Part.STEP"
    Ply_folder = r"C:\Cost_Calculation\PlyFiles"
    free_path = r"C:\Program Files\FreeCAD 0.21\bin\FreeCADCmd.exe"
    

elif desktop == False:
    colourprice = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\ColourPrice.xlsx"
    Logo = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\logo_3.png"
    res_excel = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\results3.xlsx"
    stp_folder = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\StepFiles\Case1 - Part.STEP"
    Ply_folder = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\PlyFiles"
    free_path = r"C:\Program Files\FreeCAD 0.21\bin\FreeCADCmd.exe"
    


class STEPAnalyzer(QMainWindow):
    file_selected = pyqtSignal(str)  # Signal to notify when a new file is selected

    def __init__(self):
        super().__init__()

        self.color_excel_file = colourprice
        self.is_updating_colors = False  # Guard variable to prevent recursion

        # Initialize variables with default values
        self.weight_kg = 0.0
        self.weight_g = 0.0
        self.surface_area_m2 = 0.0
        self.volume_in_m3 = 0.0
        self.volume_in_dm3 = 0.0
        self.dimensions = [0.0, 0.0, 0.0]

        # Window settings
        self.setWindowTitle("Cost Calculator")
        self.resize(700, 400)
        self.center_window()

        # Set window icon (replace with a valid path)
        self.setWindowIcon(QIcon(Logo))

        # Apply styles
        self.setStyleSheet("""
            QLabel#infoLabel {
            font-size: 16px;
            color: #333;
            padding: 5px;
            background-color: #F0F8FF;
            border: 1px solid #B0C4DE;
            border-radius: 5px;
        }
                           
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                margin-top: 20px;
                border: 2px solid #B0C4DE;
                border-radius: 25px;
                padding: 10px;
            }
            QPushButton {
                background-color: #006400;
                color: white;
                border-radius: 5px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #228B22;
            }
            QComboBox {
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Main layout
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)

        # Add the logo
        self.logo_label = QLabel()
        pixmap = QPixmap(Logo)
        self.logo_label.setPixmap(pixmap.scaledToWidth(120, Qt.SmoothTransformation))
        self.logo_label.setAlignment(Qt.AlignRight | Qt.AlignTop)
        main_layout.addWidget(self.logo_label)

        # Button for selecting a STEP file
        self.btn_select = QPushButton("Select STEP File")
        self.btn_select.clicked.connect(self.select_file)
        main_layout.addWidget(self.btn_select)

        # Layout for selected file label and material selection
        file_material_layout = QHBoxLayout()
        
        # Label for displaying selected file
        self.label_selected_file = QLabel("Selected 3D model: N/A")
        font = self.label_selected_file.font()
        self.label_selected_file.setFont(font)
        file_material_layout.addWidget(self.label_selected_file)

        # ComboBox for selecting material
        self.material_combo = QComboBox()
        self.material_combo.addItem("Steel S235JR", 7800)
        self.material_combo.addItem("Aluminum", 2700)
        self.material_combo.setFixedWidth(150)
        self.material_combo.currentIndexChanged.connect(self.update_weight)
        file_material_layout.addWidget(self.material_combo)


        # Add this to the layout next to material selection
        file_material_layout = QHBoxLayout()
        file_material_layout.addWidget(self.label_selected_file)
        file_material_layout.addWidget(self.material_combo)
        

        # Load colors from the Excel file
        main_layout.addLayout(file_material_layout)

        # Group box for analysis results
        analysis_group = QGroupBox("3D Model Analysis Results")
        main_layout.addWidget(analysis_group)

        # Grid layout for information labels and unit selectors
        info_layout = QGridLayout()
        analysis_group.setLayout(info_layout)

        # Unit selectors for weight, surface area, volume, and bounding box
        self.weight_unit_combo = QComboBox()
        self.weight_unit_combo.addItems(["kg", "g"])
        self.weight_unit_combo.currentIndexChanged.connect(self.update_display)
        self.weight_unit_combo.setFixedWidth(55)

        self.surface_area_unit_combo = QComboBox()
        self.surface_area_unit_combo.addItems(["m²", "dm²", "cm²", "mm²"])
        self.surface_area_unit_combo.currentIndexChanged.connect(self.update_display)
        self.surface_area_unit_combo.setFixedWidth(55)

        self.volume_unit_combo = QComboBox()
        self.volume_unit_combo.addItems(["m³", "dm³", "cm³", "mm³"])
        self.volume_unit_combo.currentIndexChanged.connect(self.update_display)
        self.volume_unit_combo.setFixedWidth(55)
        

        self.bounding_box_unit_combo = QComboBox()
        self.bounding_box_unit_combo.addItems(["m", "dm", "cm", "mm"])
        self.bounding_box_unit_combo.currentIndexChanged.connect(self.update_display)

        # Add unit selectors to the UI
        info_layout.addWidget(self.weight_unit_combo, 0, 3)
        info_layout.addWidget(self.surface_area_unit_combo, 1, 3)
        info_layout.addWidget(self.volume_unit_combo, 2, 3)
        info_layout.addWidget(self.bounding_box_unit_combo, 3, 3)

        
        # Information labels
        self.label_weight = QLabel("N/A")
        self.label_weight.setObjectName("infoLabel")
        self.label_surface_area = QLabel("N/A")
        self.label_surface_area.setObjectName("infoLabel")
        self.label_volume = QLabel("N/A")
        self.label_volume.setObjectName("infoLabel")
        self.label_bounding_box = QLabel("N/A")
        self.label_bounding_box.setObjectName("infoLabel")
        info_layout.addWidget(QLabel("<b>Weight:</b>"), 0, 0)
        info_layout.addWidget(self.label_weight, 0, 1)
        info_layout.addWidget(QLabel("<b>Surface Area:</b>"), 1, 0)
        info_layout.addWidget(self.label_surface_area, 1, 1)
        info_layout.addWidget(QLabel("<b>Volume:</b>"), 2, 0)
        info_layout.addWidget(self.label_volume, 2, 1)
        info_layout.addWidget(QLabel("<b>Bounding Box:</b>"), 3, 0)
        info_layout.addWidget(self.label_bounding_box, 3, 1)

        self.volume_unit_combo.setCurrentText("dm³")
        # Status label for feedback
        self.label_status = QLabel("")
        main_layout.addWidget(self.label_status)

        # Button for viewing the 3D model
        #self.btn_view_3d_model = QPushButton("Open 3D Drawing")
        #self.btn_view_3d_model.clicked.connect(self.open_3d_viewer)
        #self.btn_view_3d_model.setEnabled(False)
        #main_layout.addWidget(self.btn_view_3d_model)

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()
        self.btn_open_excel = QPushButton("Open Excel")
        self.btn_open_excel.clicked.connect(self.open_excel)
        button_layout.addWidget(self.btn_open_excel)

        self.btn_close = QPushButton("Close Program")
        self.btn_close.clicked.connect(self.close_program)
        button_layout.addWidget(self.btn_close)

        main_layout.addLayout(button_layout)

        # Paths and placeholders
        self.default_folder = os.path.dirname(stp_folder)
        self.step_file_path = None
        self.ply_folder = (Ply_folder)
        os.makedirs(self.ply_folder, exist_ok=True)
        self.excel_file = res_excel

    def center_window(self):
        screen_geometry = QApplication.primaryScreen().geometry()
        screen_width = screen_geometry.width()
        screen_height = screen_geometry.height()
        x = (screen_width - self.width()) // 2
        y = (screen_height - self.height()) // 2
        self.move(x, y)

    def select_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_path, _ = QFileDialog.getOpenFileName(self, "Select STEP File", self.default_folder, "STEP Files (*.step *.stp);;All Files (*)", options=options)
        if file_path:
            self.step_file_path = file_path
            
            # Set the .ply file path in the designated PLY folder
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            self.ply_file_path = os.path.join(self.ply_folder, f"{file_name}.ply")
            
            # Check if the PLY file already exists; if not, convert STEP to PLY
            if not os.path.exists(self.ply_file_path):
                if not self.convert_step_to_ply(self.step_file_path, self.ply_file_path):
                    self.label_status.setText("Failed to convert STEP to PLY.")
                    return  # Stop if conversion fails

            # Update label, enable view button, and emit signal
            self.label_selected_file.setText(f"Selected file: <b>{os.path.basename(file_path)}</b>")
            self.analyze_step_file(file_path)
            self.file_selected.emit(self.ply_file_path)


    def analyze_step_file(self, file_path):
        try:
            # Open and read the STEP file in FreeCAD
            doc = FreeCAD.newDocument("doc")
            shape = Part.Shape()
            shape.read(file_path)

            # Calculate surface area and store it in m²
            surface_area = shape.Area
            self.surface_area_m2 = round(surface_area / 1e6, 2)

            # Calculate volume and store it in m³
            volume = shape.Volume
            self.volume_in_m3 = volume / 1e9  # Store in m³
            self.volume_in_dm3 = volume / 1e6  # Store in dm³ for UI
            # Reset dimensions and calculate bounding box
            bounding_box = shape.BoundBox
            self.dimensions = sorted([bounding_box.XLength, bounding_box.YLength, bounding_box.ZLength], reverse=True)

            # Reset weight attributes and recalculate based on material density
            self.weight_kg = 0.0
            self.weight_g = 0.0
            self.update_weight()  # Update weight based on volume and material density

            # Check if the file is already in Excel
            self.check_excel_file(file_path)

            # Update the UI display with new values
            self.update_display()

        except Exception as e:
            self.label_status.setText(f"Error: {str(e)}")


    def check_excel_file(self, file_path):
        part_name = os.path.basename(file_path).strip()
        try:
            if os.path.exists(self.excel_file):
                existing_df = pd.read_excel(self.excel_file)
                if part_name in existing_df['Part Name'].values:
                    self.label_status.setText(f"Part '{part_name}' has already been read.")
                else:
                    self.save_to_excel(file_path)
                    self.label_status.setText(f"Data for part '{part_name}' has been saved.")
            else:
                self.save_to_excel(file_path)
                self.label_status.setText("New Excel file created and data saved.")
        except Exception as e:
            self.label_status.setText(f"Error checking Excel file: {str(e)}")

    from openpyxl.utils import get_column_letter

    def save_to_excel(self, file_path):
        part_name = os.path.basename(file_path).strip()
        # Format bounding box dimensions to 2 decimal points
        bounding_box_formatted = f"{self.dimensions[0]:.1f} x {self.dimensions[1]:.1f} x {self.dimensions[2]:.1f}"
        volume_formatted = round(self.volume_in_m3, 5)
        volume_formatted_dm3 = volume_formatted * 1000

        data = {
            'Part Name': [part_name],
            'Surface Area (m²)': [self.surface_area_m2],
            'Volume (dm³)': [volume_formatted_dm3],
            'Weight (kg)': [self.weight_kg],
            'Bounding Box (mm)': [bounding_box_formatted],
            'Date Added': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        }
        df_new = pd.DataFrame(data)

        # Load existing data or create a new DataFrame if the file doesn't exist
        if os.path.exists(self.excel_file):
            existing_df = pd.read_excel(self.excel_file)
            # Check and reorder columns if needed
            if not all(col in existing_df.columns for col in data.keys()):
                existing_df = pd.DataFrame(columns=data.keys())
            df = pd.concat([existing_df, df_new], ignore_index=True)
        else:
            df = df_new

        # Save to Excel and auto-adjust column widths
        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            # Apply auto-filter to the header row for sorting
            worksheet.auto_filter.ref = worksheet.dimensions

            # Auto-adjust column widths
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col) + 3
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max_length

            # Center align all cells except 'Part Name'
            center_alignment = Alignment(horizontal="center")
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = center_alignment



    def update_weight(self):
        density = self.material_combo.currentData()
        weight_in_kg = self.volume_in_m3 * density
        self.weight_kg = round(weight_in_kg, 2)
        self.weight_g = round(weight_in_kg * 1000, 2)
        self.update_display()

    def update_display(self):
        area_unit = self.surface_area_unit_combo.currentText()
        area_conversion = {"m²": 1, "dm²": 100, "cm²": 10000, "mm²": 1e6}
        surface_area_value = self.surface_area_m2 * area_conversion[area_unit]
        surface_area_text = f"{int(surface_area_value):.0f}" if area_unit == "mm²" else f"{surface_area_value:.2f}"
        self.label_surface_area.setText(f"{surface_area_text} {area_unit}")

        volume_unit = self.volume_unit_combo.currentText()
        volume_conversion = {"m³": 1, "dm³": 1e3, "cm³": 1e6, "mm³": 1e9}
        volume_value = self.volume_in_m3 * volume_conversion[volume_unit]
        volume_text = f"{int(volume_value):.0f}" if volume_unit == "mm³" else f"{volume_value:.3f}" if volume_unit == "m³" else f"{volume_value:.2f}"
        self.label_volume.setText(f"{volume_text} {volume_unit}")

        length_unit = self.bounding_box_unit_combo.currentText()
        length_conversion = {"m": 1, "dm": 10, "cm": 100, "mm": 1000}
        converted_dimensions = [d / 1000 * length_conversion[length_unit] for d in self.dimensions]
        bounding_box_text = " x ".join(f"{int(d):.0f}" if length_unit == "mm" else f"{d:.2f}" if length_unit == "m" else f"{d:.1f}" for d in converted_dimensions)
        self.label_bounding_box.setText(f"{bounding_box_text} {length_unit}")

        weight_unit = self.weight_unit_combo.currentText()
        self.label_weight.setText(f"{self.weight_g:.2f} g" if weight_unit == "g" else f"{self.weight_kg:.2f} kg")

    def open_excel(self):
        try:
            os.startfile(self.excel_file)
        except Exception as e:
            print(f"Error opening Excel file: {e}")

    def close_program(self):
        self.close()

    def convert_step_to_ply(self, step_path, ply_path):
        # Update this path to where FreeCADCmd.exe is installed on your system
        freecad_cmd_path = free_path
        
        if not os.path.exists(freecad_cmd_path):
            print("FreeCADCmd.exe not found. Check the FreeCAD installation path.")
            return False
        
        # FreeCAD command to convert STEP to PLY
        command = [
            freecad_cmd_path,
            "-c",
            (
                f"import Part, Mesh; "
                f"doc = Part.Shape(); "
                f"doc.read(r'{step_path}'); "
                f"mesh = Mesh.Mesh(doc.tessellate(1.0)); "
                f"mesh.write(r'{ply_path}')"
            )
        ]
        
        try:
            # Run the command and capture any output or errors
            result = subprocess.run(command, check=True, capture_output=True, text=True)
            print("Conversion output:", result.stdout)
            print("Converted", step_path, "to", ply_path)
            return True
        except subprocess.CalledProcessError as e:
            print("Failed to convert STEP to PLY:", e.stderr)
            print("Error:", e)
            return False




    def open_3d_viewer(self):
        if self.step_file_path:
            #Attempt conversion; open viewer only if conversion is successful
            if self.convert_step_to_ply(self.step_file_path, self.ply_file_path):
                self.viewer_window = QMainWindow()
                self.viewer_widget = VTKViewer(self.ply_file_path)
                self.viewer_window.setCentralWidget(self.viewer_widget)
                self.viewer_window.setWindowTitle("3D Model Viewer")
                self.viewer_window.resize(800, 600)
                self.viewer_window.show()
            else:
                print("Unable to open 3D viewer due to conversion failure.")


    def open_3d_viewer(self):
        if self.step_file_path:
            self.convert_step_to_ply(self.step_file_path, self.ply_file_path)
            self.viewer_window = QMainWindow()
            # Pass ply_file_path as the first argument, not as the parent
            self.viewer_widget = VTKViewer(self.ply_file_path, parent=self.viewer_window)
            self.viewer_window.setCentralWidget(self.viewer_widget)
            self.viewer_window.setWindowTitle("3D Model Viewer")
            self.viewer_window.resize(800, 600)
            self.viewer_window.show()


class VTKViewer(QWidget):
    def __init__(self, ply_path=None, parent=None):  # Add ply_path here
        super().__init__(parent)
        
        # Set up the VTK render window interactor
        self.vtk_widget = QVTKRenderWindowInteractor(self)
        self.vtk_widget.GetRenderWindow().SetMultiSamples(0)
        self.vtk_widget.GetRenderWindow().SetStereoTypeToCrystalEyes()

        # Layout for VTK rendering
        layout = QVBoxLayout()
        layout.addWidget(self.vtk_widget)
        layout.setContentsMargins(0, 0, 0, 0)  # Remove margins to avoid gaps
        self.setLayout(layout)

        # Set up the renderer with lime green background
        self.renderer = vtkRenderer()
        self.renderer.SetBackground(0.109, 0.193, 0.128)  # Lime green color in RGB
        self.vtk_widget.GetRenderWindow().AddRenderer(self.renderer)

        # Initialize the interactor
        self.interactor = self.vtk_widget.GetRenderWindow().GetInteractor()
        self.interactor.SetInteractorStyle(vtk.vtkInteractorStyleTrackballCamera())
        self.interactor.Initialize()

        # Load the PLY file after initialization
        if ply_path:
            self.load_ply_file(ply_path)

    def load_ply_file(self, ply_path):
        print(f"Attempting to load PLY file: {ply_path}")
        reader = vtkPLYReader()
        reader.SetFileName(ply_path)
        reader.Update()

        if reader.GetOutput().GetNumberOfPoints() == 0:
            print("PLY file is empty or not loaded correctly.")
            return

        mapper = vtkPolyDataMapper()
        mapper.SetInputConnection(reader.GetOutputPort())
        actor = vtkActor()
        actor.SetMapper(mapper)

        self.renderer.RemoveAllViewProps()
        self.renderer.AddActor(actor)
        self.renderer.ResetCamera()
        camera = self.renderer.GetActiveCamera()
        camera.Zoom(1.5)
        self.renderer.ResetCameraClippingRange()

        self.vtk_widget.GetRenderWindow().Render()
        print("PLY file loaded and rendered.")


class DataSummaryTab(QWidget):
    def __init__(self, color_excel_file, main_app):
        super().__init__()

        self.main_app = main_app  # Reference to main app to access other tabs
        self.color_excel_file = color_excel_file

        # Initialize unique colors and prices (assuming load_colors_from_excel sets these)
        self.unique_colors = []
        self.color_prices = {}
        self.load_colors_from_excel()

        # Add the color selection flag here
        self.is_color_selected = False

        # Initialize color_list early to avoid AttributeError
        self.color_list = QListWidget(self)
        self.color_list.setWindowFlags(Qt.Popup)  # Make it behave like a pop-up window
        self.color_list.setFocusPolicy(Qt.StrongFocus)  # Ensure it gets focus for navigation
        self.color_list.hide()
        self.color_list.itemClicked.connect(self.select_color)

        # Main layout for the tab
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)

        # Set consistent font size
        label_style = "font-weight: bold; font-size: 10pt;"

        # GroupBox for color selection and quantity details
        group_box = QGroupBox("Color Selection and Quantity Details")
        group_box.setStyleSheet("""
            QGroupBox {
                font-size: 12pt;
                font-weight: bold;
                color: #333;
                border: 2px solid #B0C4DE;
                border-radius: 15px;
                margin-top: 20px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                padding: 0 10px;
            }
        """)
        
        # Layout for the GroupBox content
        group_box_layout = QVBoxLayout()
        
        # First row layout for "Select Colour" and "Quantity"
        first_row_layout = QHBoxLayout()
        
        # "Select Colour:" label
        selected_colour_label = QLabel("Select Colour:")
        selected_colour_label.setStyleSheet(label_style)
        first_row_layout.addWidget(selected_colour_label)
        
        # Spacer to push "Quantity" to the right
        first_row_layout.addStretch()
        
        # "Quantity:" label and input box side-by-side
        quantity_layout = QHBoxLayout()
        quantity_label = QLabel("Quantity:")
        quantity_label.setStyleSheet(label_style)
        quantity_layout.addWidget(quantity_label)
        
        # Quantity input box
        self.quantity_input = QLineEdit()
        self.quantity_input.setPlaceholderText("Enter quantity")
        self.quantity_input.setFixedWidth(250)
        self.quantity_input.setFixedHeight(30)
        quantity_layout.addWidget(self.quantity_input)
        
        # Add the quantity layout to the first row layout
        first_row_layout.addLayout(quantity_layout)
        
        # Add the first row layout to the group box layout
        group_box_layout.addLayout(first_row_layout)


        # Second row layout for search text box and "Colour Price"
        second_row_layout = QHBoxLayout()
        
####################################### DEBUG

        # Search bar for color input aligned below "Select Colour"
        self.color_search = QLineEdit(self)
        self.color_search.setPlaceholderText("Type to search for color")
        self.color_search.textChanged.connect(self.update_color_search_results)
        self.color_search.setFixedWidth(300)
        self.color_search.setFixedHeight(30)
        second_row_layout.addWidget(self.color_search)


############################################## DEBUG



        self.quantity_input.textChanged.connect(self.on_quantity_or_color_change)
        self.color_search.textChanged.connect(self.on_quantity_or_color_change)


        # Spacer to push "Colour Price" to the right
        second_row_layout.addStretch()

        # "Colour Price:" label and price box
        self.price_label_text = QLabel("Colour Price:")
        self.price_label_text.setStyleSheet(label_style)
        second_row_layout.addWidget(self.price_label_text)

        # Price display box aligned with search box
        self.price_label = QLabel("N/A DKK/kg")
        self.price_label.setObjectName("infoLabel")
        self.price_label.setFixedWidth(222)
        self.price_label.setFixedHeight(30)
        self.price_label.setStyleSheet("""
            QLabel#infoLabel {
                font-size: 8pt;
                color: #333;
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)
        second_row_layout.addWidget(self.price_label)

        # Add the second row layout to the group box layout
        group_box_layout.addLayout(second_row_layout)

        # Add the color group box to the main layout
        group_box.setLayout(group_box_layout)
        main_layout.addWidget(group_box)

        # ==================
        # Block Selector (Radio Buttons)
        # ==================
        selector_layout = QHBoxLayout()
        self.radio_block1 = QRadioButton("Optimal Solution")
        self.radio_block2 = QRadioButton("Manual Solution")
        self.button_group = QButtonGroup()
        self.button_group.addButton(self.radio_block1)
        self.button_group.addButton(self.radio_block2)
        self.radio_block1.setChecked(True)  # Set Block 1 as enabled by default
        selector_layout.addWidget(self.radio_block1)
        selector_layout.addWidget(self.radio_block2)
        main_layout.addLayout(selector_layout)

        # Connect radio buttons to toggle function
        self.radio_block1.toggled.connect(self.toggle_blocks)
        self.radio_block2.toggled.connect(self.toggle_blocks)

        # Side-by-side Block Layout for Block 1 and Block 2
        blocks_layout = QHBoxLayout()

        # Block 1: Conveyor Space Calculation
        block1_group_box = QGroupBox("Optimal Calculator")
        block1_layout = QVBoxLayout()
        block1_group_box.setLayout(block1_layout)

        # Output labels for result display with dividers
        self.length_label = QLabel("Length of part: N/A")
        self.width_label = QLabel("Width of part: N/A")
        self.height_label = QLabel("Height of part: N/A")
        self.hanged_parts_label = QLabel("Number of hanged part(s) on one pair of hooks: N/A")
        self.result_per_part_label = QLabel("Conveyor Space per Part: N/A")
        self.result_total_space_label = QLabel("Total Conveyor Space Needed: N/A")
        
        # Adding labels with dividers
        block1_layout.addWidget(self.length_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.width_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.height_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.hanged_parts_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.result_per_part_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.result_total_space_label)

        # Add Block 1 to blocks_layout
        blocks_layout.addWidget(block1_group_box)

        # Block 2: Manual Solution
        block2_group_box = QGroupBox("Manual Calculator")
        block2_layout = QVBoxLayout()
        block2_group_box.setLayout(block2_layout)

        # Radio buttons to select hanging dimension
        self.manual_dimension_group = QButtonGroup()
        self.radio_width = QRadioButton("Hang by Width")
        self.radio_length = QRadioButton("Hang by Length")
        self.radio_height = QRadioButton("Hang by Height")

        self.manual_dimension_group.addButton(self.radio_width)
        self.manual_dimension_group.addButton(self.radio_length)
        self.manual_dimension_group.addButton(self.radio_height)

        block2_layout.addWidget(self.radio_width)
        block2_layout.addWidget(self.radio_length)
        block2_layout.addWidget(self.radio_height)

        # Connect radio buttons to trigger calculation on selection
        self.radio_width.toggled.connect(self.calculate_manual_conveyor_space)
        self.radio_length.toggled.connect(self.calculate_manual_conveyor_space)
        self.radio_height.toggled.connect(self.calculate_manual_conveyor_space)

        # Labels to display manual calculation results
        self.manual_hanged_parts_label = QLabel("Number of hanged part(s) on one pair of hooks: N/A")
        self.manual_result_per_part_label = QLabel("Conveyor Space per Part: N/A")
        self.manual_result_total_space_label = QLabel("Total Conveyor Space Needed: N/A")

        block2_layout.addWidget(self.manual_hanged_parts_label)
        block2_layout.addWidget(self.manual_result_per_part_label)
        block2_layout.addWidget(self.manual_result_total_space_label)

        # Add Block 2 to blocks_layout
        blocks_layout.addWidget(block2_group_box)

        main_layout.addLayout(blocks_layout)

        # ==================
        # Additional Block (Styled like 3D Model Analysis Results)
        # ==================
        additional_group_box = QGroupBox("Additional Information")
        additional_group_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                margin-top: 20px;
                border: 2px solid #B0C4DE;
                border-radius: 25px;
                padding: 10px;
            }
        """)
        additional_layout = QGridLayout()
        additional_group_box.setLayout(additional_layout)

        # Labels for price calculation with rectangular style
        self.total_price_label = QLabel("N/A")
        self.total_price_label.setObjectName("infoLabel")
        self.total_price_label.setFixedHeight(30)
        self.total_price_label.setStyleSheet("""
            QLabel#infoLabel {
                font-size: 8pt;
                color: #333;
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)

        self.price_per_part_label = QLabel("N/A")
        self.price_per_part_label.setObjectName("infoLabel")
        self.price_per_part_label.setFixedHeight(30)
        self.price_per_part_label.setStyleSheet("""
            QLabel#infoLabel {
                font-size: 8pt;
                color: #333;
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)

        # Add labels and their values to the grid layout
        additional_layout.addWidget(QLabel("<b>Total Price:</b>"), 0, 0)
        additional_layout.addWidget(self.total_price_label, 0, 1)
        additional_layout.addWidget(QLabel("<b>Price per Part:</b>"), 1, 0)
        additional_layout.addWidget(self.price_per_part_label, 1, 1)

        # Add the additional block to the main layout, taking the remaining space
        main_layout.addWidget(additional_group_box)

        # Set the main layout for the widget
        self.setLayout(main_layout)

        # Assign blocks to instance variables for access in toggle_blocks method
        self.block1_group_box = block1_group_box
        self.block2_group_box = block2_group_box
        self.additional_group_box = additional_group_box

        # Initialize with Block 1 enabled and Block 2 shadowed
        self.toggle_blocks()

        # Adjust sizes after defining color_list
        self.adjust_sizes()

    def toggle_blocks(self):
        """Toggle between enabling Block 1 and Block 2 based on selected radio button."""
        if self.radio_block1.isChecked():
            self.set_block_enabled(self.block1_group_box, True)
            self.set_block_enabled(self.block2_group_box, False)
        else:
            self.set_block_enabled(self.block1_group_box, False)
            self.set_block_enabled(self.block2_group_box, True)

    def set_block_enabled(self, block, enabled):
        """Enable or disable a block with shadow effect."""
        opacity_effect = QGraphicsOpacityEffect()
        opacity_effect.setOpacity(1.0 if enabled else 0.3)
        block.setGraphicsEffect(opacity_effect)
        block.setEnabled(enabled)

    def calculate_conveyor_space(self):
        """Calculate conveyor space needed based on quantity and bounding box dimensions."""
        try:
            # Retrieve quantity from DataSummaryTab's input field
            quantity = int(self.quantity_input.text())

            # Retrieve bounding box dimensions from MainApp (which fetches from Tab 1)
            bound_dim = self.main_app.get_bounding_box_dimensions()

            # Call the space calculation function
            final_space = self.space(quantity, bound_dim)

            # Display the results in labels
            self.length_label.setText(f"Length of part: {bound_dim[0]:.1f} milimetres")
            self.width_label.setText(f"Width of part: {bound_dim[1]:.1f} milimetres")
            self.height_label.setText(f"Height of part: {bound_dim[2]:.1f} milimetres")
            self.hanged_parts_label.setText(f"Number of hanged part(s) on one pair of hooks: {final_space[2]:.0f} pcs")
            self.result_per_part_label.setText(f"Conveyor Space per Part: {final_space[0]/1000:.2f} metres")
            self.result_total_space_label.setText(f"Total Conveyor Space Needed: {final_space[1]/1000:.2f} metres")

            # Always calculate total price after calculating conveyor space
            self.calculate_total_price()

        except ValueError:
            QMessageBox.warning(self, "Input Error", "Please enter valid numbers for quantity and dimensions.")


    def create_divider(self):
        """Create a thin horizontal divider line."""
        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)
        divider.setStyleSheet("color: #B0C4DE;")  # Light gray color for the divider
        return divider


    def space(self, quantity, bound_dim):
        conveyor_height = 1500  # Adjusted value (1700-200)
        bound_dim.sort(reverse=True)
        bound_dim = [int(dim) for dim in bound_dim] 
        hori_gap = bound_dim[2]
        vert_gap = bound_dim[2]

        if bound_dim[0] > 50 and bound_dim[1] <= conveyor_height and bound_dim[0] <= conveyor_height:
            # Calculate conveyor space when largest dimension is horizontal
            vert_stack1 = max(math.floor(conveyor_height / (bound_dim[1] + vert_gap)), 1)
            hori_stack1 = math.ceil(quantity / vert_stack1)
            space1 = (bound_dim[0] + hori_gap) * hori_stack1

            # Calculate conveyor space when largest dimension is vertical
            vert_stack2 = max(math.floor(conveyor_height / (bound_dim[0] + vert_gap)), 1)
            hori_stack2 = math.ceil(quantity / vert_stack2)
            space2 = (bound_dim[1] + hori_gap) * hori_stack2

            # Select the smaller space requirement
            space_final = min(space1, space2)
            if space_final == space1:
                vert_stack_final = vert_stack1
            if space_final == space2:
                vert_stack_final = vert_stack2

        elif bound_dim[0] > 50 and bound_dim[1] <= conveyor_height:
            # Calculate space only for horizontal hanging
            vert_stack1 = max(math.floor(conveyor_height / (bound_dim[1] + vert_gap)), 1)
            hori_stack1 = math.ceil(quantity / vert_stack1)
            space_final = (bound_dim[0] + hori_gap) * hori_stack1
            vert_stack_final = vert_stack1

        elif bound_dim[0] > conveyor_height and bound_dim[1] > conveyor_height:
            QMessageBox.critical(self, "Error", "Part is too big to fit on the conveyor.")
            return [0, 0]

        else:
            QMessageBox.critical(self, "Arrangement Error", "Parts should be arranged on a rack.")
            return [0, 0]

        return [space_final / quantity, space_final, vert_stack_final]
    
    def calculate_manual_conveyor_space(self):
        """Calculate conveyor space based on the selected hanging dimension in manual mode."""
        quantity = int(self.quantity_input.text())  # Assume quantity is already validated
        bound_dim = self.main_app.get_bounding_box_dimensions()  # Fetch bounding box dimensions

        # Determine selected hanging dimension
        if self.radio_width.isChecked():
            hanging_dim = bound_dim[1]  # Width
        elif self.radio_length.isChecked():
            hanging_dim = bound_dim[0]  # Length
        elif self.radio_height.isChecked():
            hanging_dim = bound_dim[2]  # Height
        else:
            return  # No dimension selected, exit the method

        # Perform manual space calculation
        final_space = self.space(quantity, bound_dim)

        # Update display labels for manual calculation
        self.manual_hanged_parts_label.setText(f"Number of hanged part(s) on one pair of hooks: {final_space[2]:.0f}")
        self.manual_result_per_part_label.setText(f"Conveyor Space per Part: {final_space[0]:.2f}")
        self.manual_result_total_space_label.setText(f"Total Conveyor Space Needed: {final_space[1]:.2f}")
        self.calculate_total_price()
    
    def calculate_total_price(self):
        try:
            # Ensure that quantity is a valid positive integer
            quantity = int(self.quantity_input.text())
            if quantity <= 0:
                raise ValueError("Quantity must be greater than 0.")

            # Get color price, ensuring it is valid
            color_name = self.color_search.text()
            if color_name in self.color_prices:
                color_price = float(self.color_prices[color_name])
            else:
                raise ValueError("Color not found in price list.")

            # Get surface area from STEPAnalyzer and ensure it is valid
            surface_area_m2 = self.main_app.step_analyzer.surface_area_m2
            if surface_area_m2 <= 0:
                raise ValueError("Surface area must be greater than 0.")

            # Determine which conveyor space to use
            conveyor_space = 0.0
            conveyor_space_text = ""
            
            if self.radio_block1.isChecked():
                conveyor_space_text = self.result_total_space_label.text()
            elif self.radio_block2.isChecked():
                conveyor_space_text = self.manual_result_total_space_label.text()

            # Use regular expression to extract the numeric value from the text
            match = re.search(r"([\d.]+)", conveyor_space_text)   
            if match:
                conveyor_space = float(match.group(1))
            else:
                raise ValueError("Invalid value for conveyor space.")

            # Calculate total price and price per part
            total_price = (surface_area_m2 * color_price * 0.2)*quantity + (conveyor_space * 136)
            price_per_part = total_price / quantity

            # Update the labels
            self.total_price_label.setText(f"{total_price:.2f} DKK")
            self.price_per_part_label.setText(f"{price_per_part:.2f} DKK")

        except ValueError as e:
            QMessageBox.warning(self, "Input Error", str(e))
            self.total_price_label.setText("Total Price: N/A DKK")
            self.price_per_part_label.setText("Price per Part: N/A DKK")


####################### DEBUG #################################
 
    def load_colors_from_excel(self):
        """Load unique colors and prices from the Excel file for color suggestions."""
        if os.path.exists(self.color_excel_file):
            df = pd.read_excel(self.color_excel_file)
            self.unique_colors = df.iloc[:, 1].dropna().unique().tolist()
            # Create a dictionary for color prices based on the second column
            self.color_prices = pd.Series(df.iloc[:, 2].values, index=df.iloc[:, 1]).to_dict()
        else:
            print(f"File not found: {self.color_excel_file}")

    def adjust_sizes(self):
        """Set search and list dimensions based on window size."""
        self.color_search.setFixedWidth(int(self.parent().width() * 0.5) if self.parent() else 400)
        self.color_list.setFixedWidth(self.color_search.width())
        self.color_list.setFixedHeight(int((self.parent().height() if self.parent() else 600) * 0.5))
        self.position_color_list()

    def resizeEvent(self, event):
        """Resize search and color list when the window is resized."""
        self.adjust_sizes()
        super().resizeEvent(event)

    def position_color_list(self):
        """Position the color list directly below the search bar."""
        search_bar_pos = self.color_search.mapTo(self, QPoint(0, self.color_search.height()))
        self.color_list.move(search_bar_pos)
        
        
    def update_color_search_results(self, text):
        """Show list of matching colors under the search field as user types."""
        self.is_color_selected = False  # Reset flag since the color might change
        self.color_list.clear()
        matching_colors = [color for color in self.unique_colors if text.lower() in color.lower()]
        for color in matching_colors:
            item = QListWidgetItem(color)
            self.color_list.addItem(item)
        if matching_colors:
            self.color_list.show()
            self.position_color_list()
        else:
            self.color_list.hide()


    def select_color(self, item):
        """Set selected color in search field, show price, and hide suggestions."""
        selected_color = item.text()
        self.color_search.setText(selected_color)
        self.color_list.hide()

        # Retrieve and display the price for the selected color
        color_price = self.color_prices.get(selected_color, "N/A")
        self.price_label.setText(f"{color_price} DKK/kg")

        # Trigger recalculation if quantity is already provided
        if self.quantity_input.text().strip():
            self.on_quantity_or_color_change()


    def on_input_change(self):
        """Trigger calculation whenever the quantity or color is changed."""
        if self.quantity_input.text().strip() and self.color_search.text().strip():
            self.calculate_conveyor_space()

    def on_quantity_or_color_change(self):
        """Calculate conveyor space and total price whenever quantity or color changes."""
        # Proceed only if both inputs are provided and valid
        if self.quantity_input.text().strip() and self.color_search.text().strip():
            try:
                # Convert quantity to integer to check validity
                quantity = int(self.quantity_input.text())
                if quantity > 0 and self.color_search.text() in self.color_prices:
                    self.calculate_conveyor_space()
            except ValueError:
                # Ignore if the input is not a valid integer yet
                pass


    def adjust_sizes(self):
        """Set search and list dimensions based on window size."""
        if hasattr(self, 'color_search') and hasattr(self, 'color_list'):
            self.color_search.setFixedWidth(int(self.parent().width() * 0.5) if self.parent() else 400)
            self.color_list.setFixedWidth(self.color_search.width())
            self.color_list.setFixedHeight(int((self.parent().height() if self.parent() else 600) * 0.5))
            self.position_color_list()

    def resizeEvent(self, event):
        """Resize search and color list when the window is resized."""
        self.adjust_sizes()
        super().resizeEvent(event)




    


    


    


####################### DEBUG #################################





# Add the MainApp class definition below your existing code
class MainApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("3D Dimension Extractor - Multi-tab Interface")
        self.resize(800, 600)

        # Initialize QTabWidget
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        # Tab 1: STEP Analyzer (already defined and set)
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, "STEP Analyzer")
        self.setup_tab1()

        # Tab 2: Data Summary with DataSummaryTab class (where Block 1 conveyor calculation is added)
        color_excel_file = colourprice  # Update with your actual path
        self.data_summary_tab = DataSummaryTab(color_excel_file, self)  # Pass reference to MainApp here
        self.tabs.addTab(self.data_summary_tab, "Data Summary")

        # Tab 3: 3D Viewer (already defined and set)
        self.tab3 = QWidget()
        self.tabs.addTab(self.tab3, "3D Viewer")
        self.setup_tab3()

    def setup_tab1(self):
        tab1_layout = QVBoxLayout()
        self.step_analyzer = STEPAnalyzer()
        tab1_layout.addWidget(self.step_analyzer)
        self.tab1.setLayout(tab1_layout)

        # Connect the signal to update the viewer in Tab 3
        self.step_analyzer.file_selected.connect(self.update_3d_viewer)

    def setup_tab2(self):
        # Layout and content for the new middle tab
        tab2_layout = QVBoxLayout()
        label = QLabel("This is the new middle tab.\nYou can add data or settings here.")
        label.setAlignment(Qt.AlignCenter)
        tab2_layout.addWidget(label)
        self.tab2.setLayout(tab2_layout)

    def setup_tab3(self):
        # Layout and content for the 3D Viewer
        tab3_layout = QVBoxLayout()
        self.viewer_widget = VTKViewer()  # Initialize without file initially
        tab3_layout.addWidget(self.viewer_widget)
        self.tab3.setLayout(tab3_layout)

    def update_3d_viewer(self, ply_file_path):
        # Load the new file in the 3D viewer widget
        self.viewer_widget.load_ply_file(ply_file_path)

    def get_bounding_box_dimensions(self):
        """
        Retrieve bounding box dimensions from STEPAnalyzer in Tab 1.
        Returns the dimensions as a list [XLength, YLength, ZLength].
        """
        # Access dimensions stored in the step_analyzer instance
        return self.step_analyzer.dimensions if self.step_analyzer else [0.0, 0.0, 0.0]





if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec_())





