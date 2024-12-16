"""  Phase 4 addons
    * add pricing for covering
    * fix spacing that is needed to bound_box[2] == spacing_vert and spacing_hori
    * add how much of paint is needed for this type of painting (make it updating which every new file)
    * add levels of quality
    * add additional expense for aluminium
    * excel file with factors 
    * maximum 4 in vert_stack
"""

import sys
import os
import math
import re
import pandas as pd
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QPushButton, QLabel, QFileDialog, QWidget, QSpacerItem, QSizePolicy,
    QHBoxLayout, QGridLayout, QGroupBox, QComboBox, QTabWidget, QLineEdit, QListWidget, QListWidgetItem, QRadioButton, QButtonGroup,
    QGraphicsOpacityEffect, QMessageBox, QFrame,
)
from PyQt5.QtGui import QIcon, QPixmap, QFontMetrics, QFont
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtCore import pyqtSignal, QTimer, QRect
from openpyxl.styles import Alignment
import FreeCAD
import Part
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import subprocess
import vtk
from vtkmodules.vtkRenderingCore import vtkRenderer, vtkActor, vtkPolyDataMapper
from vtkmodules.vtkIOPLY import vtkPLYReader
from vtkmodules.qt.QVTKRenderWindowInteractor import QVTKRenderWindowInteractor

# Specify if on desktop or laptop
desktop = True

########## Paths ##########
if desktop == True:
    colourprice = r"C:\Users\Bjørn\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\ColourPrice.xlsx"
    Logo = r"C:\Users\Bjørn\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\logo_3.png"
    res_excel = r"C:\Users\Bjørn\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\results3.xlsx"
    stp_folder = r"C:\Cost_Calculation\StepFiles\Case1 - Part.STEP"
    Ply_folder = r"C:\Cost_Calculation\PlyFiles"
    free_path = r"C:\Program Files\FreeCAD 0.21\bin\FreeCADCmd.exe"
    parameter_path = r"C:\Users\Bjørn\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\Calculation_Parameters.xlsx"
else:
    colourprice = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\ColourPrice.xlsx"
    Logo = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\logo_3.png"
    res_excel = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\results3.xlsx"
    stp_folder = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\StepFiles\Case1 - Part.STEP"
    Ply_folder = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\PlyFiles"
    free_path = r"C:\Program Files\FreeCAD 0.21\bin\FreeCADCmd.exe"
    parameter_path = r"C:\Users\mail\OneDrive - Aalborg Universitet\9. Semester\P9 Code\Dependecies\Calculation_Parameters.xlsx"


class STEPAnalyzer(QMainWindow):
    file_selected = pyqtSignal(str)  # Signal to notify when a new file is selected
    new_file_loaded = pyqtSignal()
    
    def __init__(self, main_app):
        super().__init__()
        self.main_app = main_app

        self.color_excel_file = colourprice
        self.is_updating_colors = False  # Guard variable to prevent recursion

        # Initialize variables with default values
        self.weight_kg = 0.0
        self.weight_g = 0.0
        self.main_app.surface_area_m2 = 0.0
        self.volume_in_m3 = 0.0
        self.volume_in_dm3 = 0.0
        self.main_app.dimensions = [0.0, 0.0, 0.0]
        self.main_app.fileloaded = 0

        # Window settings
        self.setWindowTitle("Cost Calculator")
        self.resize(700, 400)
        self.center_window()

        # Set window icon (replace with a valid path)
        self.setWindowIcon(QIcon(Logo))

        # Apply styles
        self.setStyleSheet("""
            QLabel#infoLabel {
            font-size: 16px;
            color: #333;
            padding: 5px;
            background-color: #F0F8FF;
            border: 1px solid #B0C4DE;
            border-radius: 5px;
        }
                           
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                margin-top: 20px;
                border: 2px solid #B0C4DE;
                border-radius: 25px;
                padding: 10px;
            }
            QPushButton {
                background-color: #006400;
                color: white;
                border-radius: 5px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #228B22;
            }
            QComboBox {
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Main layout
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)

        # Add the logo
        self.logo_label = QLabel()
        pixmap = QPixmap(Logo)
        self.logo_label.setPixmap(pixmap.scaledToWidth(120, Qt.SmoothTransformation))
        self.logo_label.setAlignment(Qt.AlignRight | Qt.AlignTop)
        main_layout.addWidget(self.logo_label)

        # Button for selecting a STEP file
        self.btn_select = QPushButton("Select STEP File")
        self.btn_select.clicked.connect(self.select_file)
        main_layout.addWidget(self.btn_select)

        # Layout for selected file label and material selection
        file_material_layout = QHBoxLayout()
        
        # Label for displaying selected file
        self.label_selected_file = QLabel("Selected 3D model: N/A")
        font = self.label_selected_file.font()
        self.label_selected_file.setFont(font)
        file_material_layout.addWidget(self.label_selected_file)

        # ComboBox for selecting material
        self.material_combo = QComboBox()
        self.material_combo.addItem("Steel S235JR", 7800)
        self.material_combo.addItem("Aluminum", 2700)
        self.material_combo.setFixedWidth(150)
        self.material_combo.currentIndexChanged.connect(self.update_weight)
        file_material_layout.addWidget(self.material_combo)


        # Add this to the layout next to material selection
        file_material_layout = QHBoxLayout()
        file_material_layout.addWidget(self.label_selected_file)
        file_material_layout.addWidget(self.material_combo)
        

        # Load colors from the Excel file
        main_layout.addLayout(file_material_layout)

        # Group box for analysis results
        analysis_group = QGroupBox("3D Model Analysis Results")
        main_layout.addWidget(analysis_group)

        # Grid layout for information labels and unit selectors
        info_layout = QGridLayout()
        analysis_group.setLayout(info_layout)

        # Unit selectors for weight, surface area, volume, and bounding box
        self.weight_unit_combo = QComboBox()
        self.weight_unit_combo.addItems(["kg", "g"])
        self.weight_unit_combo.currentIndexChanged.connect(self.update_display)
        self.weight_unit_combo.setFixedWidth(55)

        self.surface_area_unit_combo = QComboBox()
        self.surface_area_unit_combo.addItems(["m²", "dm²", "cm²", "mm²"])
        self.surface_area_unit_combo.currentIndexChanged.connect(self.update_display)
        self.surface_area_unit_combo.setFixedWidth(55)

        self.volume_unit_combo = QComboBox()
        self.volume_unit_combo.addItems(["m³", "dm³", "cm³", "mm³"])
        self.volume_unit_combo.currentIndexChanged.connect(self.update_display)
        self.volume_unit_combo.setFixedWidth(55)
        

        self.bounding_box_unit_combo = QComboBox()
        self.bounding_box_unit_combo.addItems(["m", "dm", "cm", "mm"])
        self.bounding_box_unit_combo.currentIndexChanged.connect(self.update_display)

        # Add unit selectors to the UI
        info_layout.addWidget(self.weight_unit_combo, 0, 3)
        info_layout.addWidget(self.surface_area_unit_combo, 1, 3)
        info_layout.addWidget(self.volume_unit_combo, 2, 3)
        info_layout.addWidget(self.bounding_box_unit_combo, 3, 3)

        
        # Information labels
        self.label_weight = QLabel("N/A")
        self.label_weight.setObjectName("infoLabel")
        self.label_surface_area = QLabel("N/A")
        self.label_surface_area.setObjectName("infoLabel")
        self.label_volume = QLabel("N/A")
        self.label_volume.setObjectName("infoLabel")
        self.label_bounding_box = QLabel("N/A")
        self.label_bounding_box.setObjectName("infoLabel")
        info_layout.addWidget(QLabel("<b>Weight:</b>"), 0, 0)
        info_layout.addWidget(self.label_weight, 0, 1)
        info_layout.addWidget(QLabel("<b>Surface Area:</b>"), 1, 0)
        info_layout.addWidget(self.label_surface_area, 1, 1)
        info_layout.addWidget(QLabel("<b>Volume:</b>"), 2, 0)
        info_layout.addWidget(self.label_volume, 2, 1)
        info_layout.addWidget(QLabel("<b>Bounding Box:</b>"), 3, 0)
        info_layout.addWidget(self.label_bounding_box, 3, 1)

        self.volume_unit_combo.setCurrentText("dm³")
        # Status label for feedback
        self.label_status = QLabel("")
        main_layout.addWidget(self.label_status)

        # Button for viewing the 3D model
        #self.btn_view_3d_model = QPushButton("Open 3D Drawing")
        #self.btn_view_3d_model.clicked.connect(self.open_3d_viewer)
        #self.btn_view_3d_model.setEnabled(False)
        #main_layout.addWidget(self.btn_view_3d_model)

        # Create a horizontal layout for the buttons
        button_layout = QHBoxLayout()
        self.btn_open_excel = QPushButton("Open Excel")
        self.btn_open_excel.clicked.connect(self.open_excel)
        button_layout.addWidget(self.btn_open_excel)

        self.btn_close = QPushButton("Close Program")
        self.btn_close.clicked.connect(self.close_program)
        button_layout.addWidget(self.btn_close)

        main_layout.addLayout(button_layout)

        # Paths and placeholders
        self.default_folder = os.path.dirname(stp_folder)
        self.step_file_path = None
        self.ply_folder = (Ply_folder)
        os.makedirs(self.ply_folder, exist_ok=True)
        self.excel_file = res_excel

    def center_window(self):
        screen_geometry = QApplication.primaryScreen().geometry()
        screen_width = screen_geometry.width()
        screen_height = screen_geometry.height()
        x = (screen_width - self.width()) // 2
        y = (screen_height - self.height()) // 2
        self.move(x, y)

    def select_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_path, _ = QFileDialog.getOpenFileName(self, "Select STEP File", self.default_folder, "STEP Files (*.step *.stp);;All Files (*)", options=options)
        if file_path:
            self.step_file_path = file_path
            
            # Set the .ply file path in the designated PLY folder
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            self.ply_file_path = os.path.join(self.ply_folder, f"{file_name}.ply")
            
            # Check if the PLY file already exists; if not, convert STEP to PLY
            if not os.path.exists(self.ply_file_path):
                if not self.convert_step_to_ply(self.step_file_path, self.ply_file_path):
                    self.label_status.setText("Failed to convert STEP to PLY.")
                    return  # Stop if conversion fails

            # Update label, enable view button, and emit signal
            self.label_selected_file.setText(f"Selected file: <b>{os.path.basename(file_path)}</b>")
            self.analyze_step_file(file_path)
            self.file_selected.emit(self.ply_file_path)


    def analyze_step_file(self, file_path):
        try:
            # Open and read the STEP file in FreeCAD
            doc = FreeCAD.newDocument("doc")
            shape = Part.Shape()
            shape.read(file_path)

            # Calculate surface area and store it in m²
            surface_area = shape.Area
            self.main_app.surface_area_m2 = surface_area / 1e6
            
            # Calculate volume and store it in m³
            volume = shape.Volume
            self.volume_in_m3 = volume / 1e9  # Store in m³
            self.volume_in_dm3 = volume / 1e6  # Store in dm³ for UI
            # Reset dimensions and calculate bounding box
            bounding_box = shape.BoundBox
            self.main_app.dimensions = sorted([bounding_box.XLength, bounding_box.YLength, bounding_box.ZLength], reverse=True) # In mm
            print("Bounding Box: ", bounding_box.XLength, bounding_box.YLength, bounding_box.ZLength)
            print("Volume: ", bounding_box.XLength*bounding_box.YLength*bounding_box.ZLength)
            # Reset weight attributes and recalculate based on material density
            self.weight_kg = 0.0
            self.weight_g = 0.0
            self.update_weight()  # Update weight based on volume and material density

            # Check if the file is already in Excel
            self.check_excel_file(file_path)

            # Update the UI display with new values
            self.update_display()
            self.main_app.fileloaded = 1
            self.new_file_loaded.emit()

        except Exception as e:
            self.label_status.setText(f"Error: {str(e)}")


    def check_excel_file(self, file_path):
        part_name = os.path.basename(file_path).strip()
        try:
            if os.path.exists(self.excel_file):
                existing_df = pd.read_excel(self.excel_file)
                if part_name in existing_df['Part Name'].values:
                    self.label_status.setText(f"Part '{part_name}' has already been read.")
                else:
                    self.save_to_excel(file_path)
                    self.label_status.setText(f"Data for part '{part_name}' has been saved.")
            else:
                self.save_to_excel(file_path)
                self.label_status.setText("New Excel file created and data saved.")
        except Exception as e:
            self.label_status.setText(f"Error checking Excel file: {str(e)}")

    from openpyxl.utils import get_column_letter

    def save_to_excel(self, file_path):
        part_name = os.path.basename(file_path).strip()
        # Format bounding box dimensions to 2 decimal points
        bounding_box_formatted = f"{self.dimensions[0]:.1f} x {self.dimensions[1]:.1f} x {self.dimensions[2]:.1f}"
        volume_formatted = round(self.volume_in_m3, 5)
        volume_formatted_dm3 = volume_formatted * 1000

        data = {
            'Part Name': [part_name],
            'Surface Area (m²)': [self.surface_area_m2],
            'Volume (dm³)': [volume_formatted_dm3],
            'Weight (kg)': [self.weight_kg],
            'Bounding Box (mm)': [bounding_box_formatted],
            'Date Added': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        }
        df_new = pd.DataFrame(data)

        # Load existing data or create a new DataFrame if the file doesn't exist
        if os.path.exists(self.excel_file):
            existing_df = pd.read_excel(self.excel_file)
            # Check and reorder columns if needed
            if not all(col in existing_df.columns for col in data.keys()):
                existing_df = pd.DataFrame(columns=data.keys())
            df = pd.concat([existing_df, df_new], ignore_index=True)
        else:
            df = df_new

        # Save to Excel and auto-adjust column widths
        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            # Apply auto-filter to the header row for sorting
            worksheet.auto_filter.ref = worksheet.dimensions

            # Auto-adjust column widths
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col) + 3
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max_length

            # Center align all cells except 'Part Name'
            center_alignment = Alignment(horizontal="center")
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = center_alignment



    def update_weight(self):
        density = self.material_combo.currentData()
        weight_in_kg = self.volume_in_m3 * density
        self.weight_kg = round(weight_in_kg, 2)
        self.weight_g = round(weight_in_kg * 1000, 2)
        self.update_display()

    def update_display(self):
        area_unit = self.surface_area_unit_combo.currentText()
        area_conversion = {"m²": 1, "dm²": 100, "cm²": 10000, "mm²": 1e6}
        surface_area_value = round(self.main_app.surface_area_m2 * area_conversion[area_unit], 5)
        surface_area_text = f"{int(surface_area_value):.4f}" if area_unit == "mm²" else f"{surface_area_value:.4f}"
        self.label_surface_area.setText(f"{surface_area_text} {area_unit}")

        volume_unit = self.volume_unit_combo.currentText()
        volume_conversion = {"m³": 1, "dm³": 1e3, "cm³": 1e6, "mm³": 1e9}
        volume_value = self.volume_in_m3 * volume_conversion[volume_unit]
        volume_text = f"{int(volume_value):.4f}" if volume_unit == "mm³" else f"{volume_value:.4f}" if volume_unit == "m³" else f"{volume_value:.4f}"
        self.label_volume.setText(f"{volume_text} {volume_unit}")

        length_unit = self.bounding_box_unit_combo.currentText()
        length_conversion = {"m": 1, "dm": 10, "cm": 100, "mm": 1000}
        converted_dimensions = [d / 1000 * length_conversion[length_unit] for d in self.main_app.dimensions]
        bounding_box_text = " x ".join(f"{int(d):.4f}" if length_unit == "mm" else f"{d:.4f}" if length_unit == "m" else f"{d:.4f}" for d in converted_dimensions)
        self.label_bounding_box.setText(f"{bounding_box_text} {length_unit}")

        weight_unit = self.weight_unit_combo.currentText()
        self.label_weight.setText(f"{self.weight_g:.4f} g" if weight_unit == "g" else f"{self.weight_kg:.4f} kg")

    def open_excel(self):
        try:
            os.startfile(self.excel_file)
        except Exception as e:
            print(f"Error opening Excel file: {e}")

    def close_program(self):
        self.close()

    def convert_step_to_ply(self, step_path, ply_path):
        # Update this path to where FreeCADCmd.exe is installed on your system
        freecad_cmd_path = free_path
        
        if not os.path.exists(freecad_cmd_path):
            print("FreeCADCmd.exe not found. Check the FreeCAD installation path.")
            return False
        
        # FreeCAD command to convert STEP to PLY
        command = [
            freecad_cmd_path,
            "-c",
            (
                f"import Part, Mesh; "
                f"doc = Part.Shape(); "
                f"doc.read(r'{step_path}'); "
                f"mesh = Mesh.Mesh(doc.tessellate(1.0)); "
                f"mesh.write(r'{ply_path}')"
            )
        ]
        
        try:
            # Run the command and capture any output or errors
            result = subprocess.run(command, check=True, capture_output=True, text=True)
            print("Conversion output:", result.stdout)
            print("Converted", step_path, "to", ply_path)
            return True
        except subprocess.CalledProcessError as e:
            print("Failed to convert STEP to PLY:", e.stderr)
            print("Error:", e)
            return False




    def open_3d_viewer(self):
        if self.step_file_path:
            #Attempt conversion; open viewer only if conversion is successful
            if self.convert_step_to_ply(self.step_file_path, self.ply_file_path):
                self.viewer_window = QMainWindow()
                self.viewer_widget = VTKViewer(self.ply_file_path)
                self.viewer_window.setCentralWidget(self.viewer_widget)
                self.viewer_window.setWindowTitle("3D Model Viewer")
                self.viewer_window.resize(800, 600)
                self.viewer_window.show()
            else:
                print("Unable to open 3D viewer due to conversion failure.")


    def open_3d_viewer(self):
        if self.step_file_path:
            self.convert_step_to_ply(self.step_file_path, self.ply_file_path)
            self.viewer_window = QMainWindow()
            # Pass ply_file_path as the first argument, not as the parent
            self.viewer_widget = VTKViewer(self.ply_file_path, parent=self.viewer_window)
            self.viewer_window.setCentralWidget(self.viewer_widget)
            self.viewer_window.setWindowTitle("3D Model Viewer")
            self.viewer_window.resize(800, 600)
            self.viewer_window.show()

class VTKViewer(QWidget):
    def __init__(self, ply_path=None, parent=None):  # Add ply_path here
        super().__init__(parent)
        
        # Set up the VTK render window interactor
        self.vtk_widget = QVTKRenderWindowInteractor(self)
        self.vtk_widget.GetRenderWindow().SetMultiSamples(0)
        self.vtk_widget.GetRenderWindow().SetStereoTypeToCrystalEyes()

        # Layout for VTK rendering
        layout = QVBoxLayout()
        layout.addWidget(self.vtk_widget)
        layout.setContentsMargins(0, 0, 0, 0)  # Remove margins to avoid gaps
        self.setLayout(layout)

        # Set up the renderer with lime green background
        self.renderer = vtkRenderer()
        self.renderer.SetBackground(0.109, 0.193, 0.128)  # Lime green color in RGB
        self.vtk_widget.GetRenderWindow().AddRenderer(self.renderer)

        # Initialize the interactor
        self.interactor = self.vtk_widget.GetRenderWindow().GetInteractor()
        self.interactor.SetInteractorStyle(vtk.vtkInteractorStyleTrackballCamera())
        self.interactor.Initialize()

        # Load the PLY file after initialization
        if ply_path:
            self.load_ply_file(ply_path)

    def load_ply_file(self, ply_path):
        print(f"Attempting to load PLY file: {ply_path}")
        reader = vtkPLYReader()
        reader.SetFileName(ply_path)
        reader.Update()

        if reader.GetOutput().GetNumberOfPoints() == 0:
            print("PLY file is empty or not loaded correctly.")
            return

        mapper = vtkPolyDataMapper()
        mapper.SetInputConnection(reader.GetOutputPort())
        actor = vtkActor()
        actor.SetMapper(mapper)

        self.renderer.RemoveAllViewProps()
        self.renderer.AddActor(actor)
        self.renderer.ResetCamera()
        camera = self.renderer.GetActiveCamera()
        camera.Zoom(1.5)
        self.renderer.ResetCameraClippingRange()

        self.vtk_widget.GetRenderWindow().Render()
        print("PLY file loaded and rendered.")

class Calculation(QWidget):
    calc_space_signal = pyqtSignal()
    

    def __init__(self, colour_file_path, parameter_path, main_app):
        self.quant_change = 0
        super().__init__()

        self.main_app = main_app  # Allows access to other tabs
        self.colour_file_path = colour_file_path
        self.part_hanging = PartHanging(self.main_app)
    
        # Load Calculation Parameters from excel
        self.parameter_path = parameter_path
        self.calc_parameter = self.load_calculation_parameters_from_excel()
        print(self.calc_parameter)
        

        # Main layout for the tab
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)

        # Set consistent font size
        label_style = "font-weight: bold; font-size: 10pt;"
        label_style2 = "font-weight: bold; font-size: 8pt;"

        # GroupBox for color selection and quantity details
        group_box = QGroupBox("Color Selection and Quantity Details")
        group_box.setStyleSheet("""
            QGroupBox {
                font-size: 12pt;
                font-weight: bold;
                color: #333;
                border: 2px solid #B0C4DE;
                border-radius: 15px;
                margin-top: 20px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                padding: 0 10px;
            }
        """)

        # Layout for the Groupbox content
        group_box_layout = QVBoxLayout()

        # Row layout
        first_row_layout = QHBoxLayout()
        second_row_layout = QHBoxLayout()


        # Load colors from excel
        self.colour_prices = self.load_colours_from_excel()
        self.selected_price = 1  # Variable to hold price for selected color

        # Create combo box
        self.combo_box = QComboBox(self)
        self.set_combo_box_width()  # Set fixed width based on the longest item

        # Update combo box with colors
        self.update_combo_box(self.colour_prices)

        # Create the search box
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Search for colour")

        # Connect search box and combo box
        self.search_box.textChanged.connect(self.filter_colours)
        self.combo_box.currentIndexChanged.connect(self.update_selected_price)

        # Add widgets
        first_row_layout.addWidget(self.search_box)
        first_row_layout.addWidget(self.combo_box)
        group_box_layout.addLayout(first_row_layout)
        group_box.setLayout(group_box_layout)



        # "Quantity:" label and input box side-by-side
        quantity_layout = QHBoxLayout()
        quantity_label = QLabel("Quantity:")
        quantity_label.setStyleSheet(label_style)
        quantity_layout.addWidget(quantity_label)

        # Quantity input box
        self.quantity_input = QLineEdit()
        self.quantity_input.setPlaceholderText("Enter quantity")
        self.quantity_input.setFixedWidth(250)
        self.quantity_input.setFixedHeight(30)
        # Connect quantity input to calc_powder_use
        self.quantity_input.textChanged.connect(self.quantity_changed)
        self.quantity_input.textChanged.connect(self.calc_powder_use)
        #self.quantity_input.textChanged.connect(self.calculate_price)
        quantity_layout.addWidget(self.quantity_input, alignment=Qt.AlignLeft)


        # Add quantity to layout
        second_row_layout.addLayout(quantity_layout)
        group_box_layout.addLayout(second_row_layout)


        # Powder price box
        self.powdercost_label = QLabel("Powder Cost:")
        self.powdercost_label.setStyleSheet(label_style)
        second_row_layout.addWidget(self.powdercost_label)

        self.powdercost = QLabel("N/A DKK/kg")
        self.powdercost.setObjectName("infoLabel")
        self.powdercost.setFixedWidth(100)
        self.powdercost.setFixedHeight(30)
        self.powdercost.setStyleSheet("""
            QLabel#infoLabel {
                font-size: 8pt;
                color: #333;
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)
        second_row_layout.addWidget(self.powdercost)

        
        # Total powder used for order
        self.Powder_use_label = QLabel("Powder usage:")
        self.Powder_use_label.setStyleSheet(label_style)
        second_row_layout.addWidget(self.Powder_use_label)

        self.powder_use = QLabel("N/A kg")
        self.powder_use.setObjectName("infoLabel")
        self.powder_use.setFixedWidth(100)
        self.powder_use.setFixedHeight(30)
        self.powder_use.setStyleSheet("""
            QLabel#infoLabel {
                font-size: 8pt;
                color: #333;
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)

        second_row_layout.addWidget(self.powder_use)



        # Add groupbox for quality and coverings
        quality_cover_box = QGroupBox("Quality and Coverings")
        quality_cover_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                margin-top: 20px;
                border: 2px solid #B0C4DE;
                border-radius: 25px;
                padding: 10px;
            }
        """)

        quality_cover_layout = QGridLayout()
        quality_cover_box.setLayout(quality_cover_layout)

        # Label for quality level
        quality_lvl_label = QLabel("Select quality level")
        quality_lvl_label.setStyleSheet(label_style2)
        quality_lvl_label.setAlignment(Qt.AlignHCenter)

        # Add quality level label to span across two columns on the left
        quality_cover_layout.addWidget(quality_lvl_label, 0, 0, 1, 2, Qt.AlignCenter)

        # Buttons for quality
        self.quality_lvl1 = QRadioButton("Quality level 1")
        self.quality_lvl2 = QRadioButton("Quality level 2")
        self.quality_lvl3 = QRadioButton("Quality level 3")
        self.quality_lvl4 = QRadioButton("Quality level 4")

        # Create a QButtonGroup for quality-related buttons
        self.quality_button_group = QButtonGroup()
        self.quality_button_group.addButton(self.quality_lvl1)
        self.quality_button_group.addButton(self.quality_lvl2)
        self.quality_button_group.addButton(self.quality_lvl3)
        self.quality_button_group.addButton(self.quality_lvl4)

        # Reduce margins around buttons for tighter spacing
        self.quality_lvl1.setStyleSheet("margin-right: 5px;")
        self.quality_lvl2.setStyleSheet("margin-right: 5px;")
        self.quality_lvl3.setStyleSheet("margin-left: 5px;")
        self.quality_lvl4.setStyleSheet("margin-left: 5px;")

        # Arrange buttons in a grid on the left
        quality_cover_layout.addWidget(self.quality_lvl1, 1, 0, Qt.AlignRight)
        quality_cover_layout.addWidget(self.quality_lvl2, 2, 0, Qt.AlignRight)
        quality_cover_layout.addWidget(self.quality_lvl3, 1, 1, Qt.AlignLeft)
        quality_cover_layout.addWidget(self.quality_lvl4, 2, 1, Qt.AlignLeft)

        # Add horizontal stretch between quality and coverings
        quality_cover_layout.setColumnStretch(2, 1)

        # Label for coverings
        covering_label = QLabel("Number of coverings per part")
        covering_label.setStyleSheet(label_style2)

        # Place covering label on the right
        quality_cover_layout.addWidget(covering_label, 0, 3, 1, 1, Qt.AlignLeft)

        # Input for coverings
        self.covering_input = QLineEdit()
        self.covering_input.setPlaceholderText("Enter No. of coverings")
        self.covering_input.setFixedWidth(150)
        self.covering_input.setFixedHeight(30)

        # Place covering input on the right
        quality_cover_layout.addWidget(self.covering_input, 1, 3, 2, 1, Qt.AlignLeft)

        # Add horizontal stretch between coverings and material
        quality_cover_layout.setColumnStretch(4, 1)

        # Label for material selection
        material_label = QLabel("Select Material")
        material_label.setStyleSheet(label_style2)
        material_label.setAlignment(Qt.AlignHCenter)

        # Add material label in a new column (right-most)
        quality_cover_layout.addWidget(material_label, 0, 5, Qt.AlignCenter)

        # Buttons for material selection
        self.steel_button = QRadioButton("Steel")
        self.aluminium_button = QRadioButton("Aluminium")

        # Create a QButtonGroup for material-related buttons
        self.material_button_group = QButtonGroup()
        self.material_button_group.addButton(self.steel_button)
        self.material_button_group.addButton(self.aluminium_button)

        # Arrange material buttons below the label
        quality_cover_layout.addWidget(self.steel_button, 1, 5, Qt.AlignLeft)
        quality_cover_layout.addWidget(self.aluminium_button, 2, 5, Qt.AlignLeft)

        # Add stretch at the end for symmetry
        quality_cover_layout.setColumnStretch(6, 1)


        self.quality_lvl3.setChecked(True)
        self.steel_button.setChecked(True)
        
        

        self.material_button_group.buttonToggled.connect(self.calculate_price)
        self.quality_button_group.buttonToggled.connect(self.calculate_price)
        self.covering_input.textChanged.connect(self.calculate_price)



        # Add groupbox for showing the calculated price
        price_box = QGroupBox("Calculated Price")
        price_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                margin-top: 20px;
                border: 2px solid #B0C4DE;
                border-radius: 25px;
                padding: 10px;
            }
        """)
        
        additional_layout = QGridLayout()
        price_box.setLayout(additional_layout)

        self.total_price = QLabel("N/A")
        self.total_price.setObjectName("infoLabel")
        self.total_price.setFixedHeight(30)
        self.total_price.setStyleSheet("""
            QLabel#infoLabel {
                font-size: 8pt;
                color: #333;
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)

        self.price_per_part = QLabel("N/A")
        self.price_per_part.setObjectName("infoLabel")
        self.price_per_part.setFixedHeight(30)
        self.price_per_part.setStyleSheet("""
            QLabel#infoLabel {
                font-size: 8pt;
                color: #333;
                padding: 5px;
                background-color: #F0F8FF;
                border: 1px solid #B0C4DE;
                border-radius: 5px;
            }
        """)

        additional_layout.addWidget(QLabel("<b>Total Price:</b>"), 0, 0)
        additional_layout.addWidget(self.total_price, 0, 1)
        additional_layout.addWidget(QLabel("<b>Price per Part:</b>"), 1, 0)
        additional_layout.addWidget(self.price_per_part, 1, 1)

       

        main_layout.addWidget(group_box)
        main_layout.addWidget(quality_cover_box)
        main_layout.addWidget(price_box)
        self.setLayout(main_layout)


    def load_calculation_parameters_from_excel(self):
        try:
            df = pd.read_excel(self.parameter_path)
            calc_parameter = pd.Series(df.Value.values, index=df.Parameter).to_dict()
            return calc_parameter
        except Exception as e:
            print(F"Error loading excel file: {e}")
            return None

    def load_colours_from_excel(self):
        try:
            df = pd.read_excel(self.colour_file_path)
            colourprices = pd.Series(df.Price.values, index=df.Colour).to_dict()
            return colourprices
        except Exception as e:
            print(f"Error loading excel file: {e}")
            return {}

    def set_combo_box_width(self):
        # Calculate the width based on the longest item in the full list
        longest_text = ""
        for colour, price in self.colour_prices.items():
            display_text = f"{colour} - ${price:.2f}"
            if len(display_text) > len(longest_text):
                longest_text = display_text

        # Calculate the pixel width of the longest text
        font_metrics = QFontMetrics(self.combo_box.font())
        text_width = font_metrics.horizontalAdvance(longest_text) + 20  # Add padding

        # Set the width of the combo box to fit the longest item
        self.combo_box.setFixedWidth(text_width)

    def filter_colours(self):
        filter_text = self.search_box.text().lower()

        # Filter items based on the search box
        filtered_items = {colour: price for colour, price in self.colour_prices.items() if filter_text in colour.lower()}

        # Temporarily disconnect the signal to avoid recursion while updating the combo box
        self.combo_box.currentIndexChanged.disconnect(self.update_selected_price)

        # Update combo box with filtered items
        self.update_combo_box(filtered_items)

        # Reconnect signal to combo box
        self.combo_box.currentIndexChanged.connect(self.update_selected_price)

    def update_combo_box(self, items):
        self.combo_box.clear()

        # Style adjustments for readability
        self.combo_box.setStyleSheet("""
            QComboBox QAbstractItemView {
                padding: 5px;
                font-size: 10pt;
                background-color: #F0F0F0;
                selection-background-color: #B0C4DE;
            }
            QComboBox {
                padding: 5px;
                font-size: 10pt;
            }
        """)

        # Add items to the combo box
        for colour, price in items.items():
            display_text = f"{colour}"
            self.combo_box.addItem(display_text, price)

    def update_selected_price(self):
        # Retrieve the price directly based on combo box selection
        self.selected_price = self.combo_box.currentData()
        self.powdercost.setText(f"{self.selected_price} DKK/kg")
        self.calculate_price()


    def calc_powder_use(self):
        try:
            quantity = int(self.quantity_input.text())
            powder_usage = round((quantity * self.main_app.surface_area_m2)/5, 2)
            self.powder_use.setText(f"{powder_usage} kg")

        except ValueError:
            self.powder_use.setText("N/A kg")


    def quantity_changed(self):
        try:
            quantity = int(self.quantity_input.text())
            if quantity > 0:
                self.quant_change = 1
                self.calculate_price()
            else:
                self.quant_change = 0

        except ValueError:
            self.quant_change = 0
            pass


    def covering(self):
        try:
            covering = int(self.covering_input.text())
            return covering
        except ValueError:
            covering = 0
            return covering

    def quality(self):
        if self.quality_lvl1.isChecked():
            quality = 2 * self.calc_parameter["QualityLevel"]/100
        elif self.quality_lvl2.isChecked():
            quality = self.calc_parameter["QualityLevel"]/100
        elif self.quality_lvl3.isChecked():
            quality = 0
        elif self.quality_lvl4.isChecked():
            quality = -self.calc_parameter["QualityLevel"]/100
        return quality

    def material(self):
        if self.steel_button.isChecked():
            material = 0
        elif self.aluminium_button.isChecked():
            material = self.calc_parameter["Aluminium"]/100
        return material


    def calculate_price(self):

        if self.quant_change == 1 and self.main_app.fileloaded == 1:
            self.main_app.quantity = int(self.quantity_input.text())
            quantity = int(self.quantity_input.text())

            self.calc_space_signal.emit()
            conveyor_space = self.main_app.space_final
            print(conveyor_space)

            covering = self.covering() * self.calc_parameter["CoveringCost"]
            quality = self.quality()
            print(f"quality: {quality}")
            material = self.material()
            profit = 1 + self.calc_parameter["ProfitMargin"]/100

            calc_total_price = ((self.main_app.surface_area_m2 * self.selected_price * 0.2 + covering)*quantity + (conveyor_space * self.calc_parameter["Overhead"]) * (1 + quality + material)) * profit
            #calc_total_price = (((self.main_app.surface_area_m2 * self.selected_price * 0.2 + covering)*quantity + (conveyor_space * material * self.calc_parameter["Overhead"])) * quality) * profit
            calc_part_price = calc_total_price/quantity

            self.total_price.setText(f"{round(calc_total_price, 2)} DKK")
            self.price_per_part.setText(f"{round(calc_part_price, 2)} DKK")

        else:
            print("More inputs needed!!!")

class PartHanging(QWidget):
    space_update = pyqtSignal()

    number_parts = 0
    conveyor_space_part = 0

    def __init__(self, main_app):
        super().__init__()

        self.main_app = main_app

        main_layout = QVBoxLayout()
        #main_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)

        # set consistent label style
        label_style = "font-weight: bold; font-size: 10pt;"

        selector_layout = QHBoxLayout()
        self.radio_block1 = QRadioButton("Optimal Solution")
        self.radio_block2 = QRadioButton("Manual Solution")
        self.button_group = QButtonGroup()
        self.button_group.addButton(self.radio_block1)
        self.button_group.addButton(self.radio_block2)
        self.radio_block1.setChecked(True)  # Set Block 1 as enabled by default
        selector_layout.addWidget(self.radio_block1)
        selector_layout.addWidget(self.radio_block2)
        main_layout.addLayout(selector_layout)

        # Connect radio buttons to toggle function
        self.radio_block1.toggled.connect(self.toggle_blocks)
        self.radio_block2.toggled.connect(self.toggle_blocks)

        # Side-by-side Block Layout for Block 1 and Block 2
        blocks_layout = QHBoxLayout()

        # Block 1: Conveyor Space Calculation
        block1_group_box = QGroupBox("Optimal Calculator")
        block1_layout = QVBoxLayout()
        block1_group_box.setLayout(block1_layout)

        # Output labels for result display with dividers
        self.length_label = QLabel("Length of part: N/A")
        self.width_label = QLabel("Width of part: N/A")
        self.height_label = QLabel("Height of part: N/A")
        self.hanged_parts_label = QLabel("Number of hanged parts on one pair of hooks: N/A")
        self.result_per_part_label = QLabel("Conveyor Space per Part: N/A")
        self.result_total_space_label = QLabel("Total Conveyor Space Needed: N/A")

        # Adding labels with dividers
        block1_layout.addWidget(self.length_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.width_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.height_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.hanged_parts_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.result_per_part_label)
        block1_layout.addWidget(self.create_divider())
        block1_layout.addWidget(self.result_total_space_label)

        # Add Block 1 to blocks_layout
        blocks_layout.addWidget(block1_group_box)

        # Block 2: Manual Solution
        block2_group_box = QGroupBox("Manual Calculator")
        block2_layout = QVBoxLayout()
        block2_group_box.setLayout(block2_layout)



        # Inputs for the manual hanging of parts
        block2_group_box_input = QGroupBox()
        block2_input_layout = QGridLayout()
        block2_group_box_input.setLayout(block2_input_layout)

        # Labels for the inputs
        number_parts_label = QLabel("No. of parts:")
        number_parts_label.setStyleSheet(label_style)
        conveyor_space_label = QLabel("Conveyor space [m]:")
        conveyor_space_label.setStyleSheet(label_style)
        # Add labels to the layout
        block2_input_layout.addWidget(number_parts_label, 1, 1)
        block2_input_layout.addWidget(conveyor_space_label, 1, 2)

        # Inputs
        self.number_parts_input = QLineEdit()
        self.number_parts_input.setPlaceholderText("Enter number of parts")
        self.number_parts_input.setFixedWidth(150)
        self.number_parts_input.setFixedHeight(30)
        self.number_parts_input.textChanged.connect(self.space_update_fun)
        self.conveyor_space_input = QLineEdit()
        self.conveyor_space_input.setPlaceholderText("Enter conveyor space")
        self.conveyor_space_input.setFixedWidth(150)
        self.conveyor_space_input.setFixedHeight(30)
        self.conveyor_space_input.textChanged.connect(self.space_update_fun)
        # Add inputs to the layout
        block2_input_layout.addWidget(self.number_parts_input, 2, 1)
        block2_input_layout.addWidget(self.conveyor_space_input, 2, 2)




        # Add the input layout to the manual block layout
        block2_layout.addWidget(block2_group_box_input)

        # Labels to display manual calculation results
        self.manual_hanged_parts_label = QLabel("Number of hanged parts on one pair of hooks: N/A")
        self.manual_result_per_part_label = QLabel("Conveyor Space per Part: N/A")
        self.manual_result_total_space_label = QLabel("Total Conveyor Space Needed: N/A")

        block2_layout.addWidget(self.manual_hanged_parts_label)
        block2_layout.addWidget(self.manual_result_per_part_label)
        block2_layout.addWidget(self.manual_result_total_space_label)

        # Add Block 2 to blocks_layout
        blocks_layout.addWidget(block2_group_box)

        main_layout.addLayout(blocks_layout)


        # Assign blocks to instance variables for access in toggle_blocks method
        self.block1_group_box = block1_group_box
        self.block2_group_box = block2_group_box
        

        # Initialize with Block 1 enabled and Block 2 shadowed
        self.toggle_blocks()

        self.setLayout(main_layout)


    def toggle_blocks(self):
        """Toggle between enabling Block 1 and Block 2 based on selected radio button."""
        if self.radio_block1.isChecked():
            self.set_block_enabled(self.block1_group_box, True)
            self.set_block_enabled(self.block2_group_box, False)
            self.main_app.optimal_manual = True
            print(self.main_app.optimal_manual)
            self.space_update.emit()

        else:
            self.set_block_enabled(self.block1_group_box, False)
            self.set_block_enabled(self.block2_group_box, True)
            self.main_app.optimal_manual = False
            print(self.main_app.optimal_manual)
            self.space_update.emit()


    def set_block_enabled(self, block, enabled):
        """Enable or disable a block with shadow effect."""
        opacity_effect = QGraphicsOpacityEffect()
        opacity_effect.setOpacity(1.0 if enabled else 0.3)
        block.setGraphicsEffect(opacity_effect)
        block.setEnabled(enabled)

    def create_divider(self):
        """Create a thin horizontal divider line."""
        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)
        divider.setStyleSheet("color: #B0C4DE;")  # Light gray color for the divider
        return divider

    def space_update_fun(self):
        self.space_update.emit()

    def space(self):
        if self.main_app.optimal_manual == True:
            self.optimal_space()
        elif self.main_app.optimal_manual == False:
            self.manual_space()


    def manual_space(self):
        try:
            number_parts = int(self.number_parts_input.text())
            conveyor_space_part = float(self.conveyor_space_input.text())
            quantity = self.main_app.quantity
 
            space_final = (quantity/number_parts) * (conveyor_space_part + 0.05)

            self.manual_hanged_parts_label.setText(f"Number of hanged parts on one pair of hooks: {number_parts}")
            self.manual_result_per_part_label.setText(f"Conveyor Space per Part: {round(space_final/quantity, 2)}")
            self.manual_result_total_space_label.setText(f"Total Conveyor Space Needed: {round(space_final, 2)}")

            self.main_app.space_final = space_final
            
        except ValueError:
            #QMessageBox.critical(self, "Error", "Invalid input for number of parts or conveyor space")
            pass


    def optimal_space(self):
        bound_dim = self.main_app.dimensions
        quantity = self.main_app.quantity
        hori_gap = bound_dim[2] + 50
        vert_gap = bound_dim[2] + 50
        conveyor_height = 1500 # 1700-200

        if bound_dim[0] > 20 and bound_dim[1] <= conveyor_height and bound_dim[0] <= conveyor_height:
            # Calculate conveyor space when largest dimension is horizontal
            vert_stack1 = max(math.floor(conveyor_height / (bound_dim[1] + vert_gap)), 1)
            vert_stack1 = min(vert_stack1, 4)
            hori_stack1 = math.ceil(quantity / vert_stack1)
            space1 = (bound_dim[0] + hori_gap) * hori_stack1

            # Calculate conveyor space when largest dimension is vertical
            vert_stack2 = max(math.floor(conveyor_height / (bound_dim[0] + vert_gap)), 1)
            vert_stack2 = min(vert_stack2, 4)
            hori_stack2 = math.ceil(quantity / vert_stack2)
            space2 = (bound_dim[1] + hori_gap) * hori_stack2

            # Select the smaller space requirement
            space_final = min(space1, space2)
            if space_final == space1:
                vert_stack_final = vert_stack1
                self.length_label.setText(f"Length of part: {round(bound_dim[0], 2)} mm")
                self.height_label.setText(f"Height of part: {round(bound_dim[1], 2)} mm")
            if space_final == space2:
                vert_stack_final = vert_stack2
                self.length_label.setText(f"Length of part: {round(bound_dim[1], 2)} mm")
                self.height_label.setText(f"Height of part: {round(bound_dim[0], 2)} mm")

        elif bound_dim[0] > 20 and bound_dim[1] <= conveyor_height:
            # Calculate space only for horizontal hanging
            vert_stack1 = max(math.floor(conveyor_height / (bound_dim[1] + vert_gap)), 1)
            vert_stack1 = min(vert_stack1, 4)
            hori_stack1 = math.ceil(quantity / vert_stack1)
            space_final = (bound_dim[0] + hori_gap) * hori_stack1
            vert_stack_final = vert_stack1

        elif bound_dim[0] > conveyor_height and bound_dim[1] > conveyor_height:
            QMessageBox.critical(self, "Error", "Part is too big to fit on the conveyor.")
            

        else:
            QMessageBox.critical(self, "Arrangement Error", "Parts should be arranged on rack.")
            


        bound_dim[2] = round(bound_dim[2], 2)

        space_final = round(space_final, 2)


        self.width_label.setText(f"Width of part: {bound_dim[2]} mm")
        self.hanged_parts_label.setText(f"Number of hanged part(s) on one pair of hooks: {vert_stack_final} parts")
        self.result_per_part_label.setText(f"Conveyor Space per Part: {round((space_final/quantity)/1000, 2)} m")
        self.result_total_space_label.setText(f"Total Conveyor Space Needed: {round(space_final/1000, 2)} m")
        
        self.main_app.space_final = space_final/1000
        self.main_app.vert_stack_final = vert_stack_final

class MainApp(QMainWindow):
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("3D Dimension Extractor - Multi-tab Interface")
        self.resize(800, 600)
 
        # Initialize QTabWidget
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        self.surface_area_m2 = 1

        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, "STEP Analyzer")
        self.setup_tab1()

        self.tab2 = QWidget()
        self.tabs.addTab(self.tab2, "Calculation")
        self.setup_tab2()

        self.tab3 = QWidget()
        self.tabs.addTab(self.tab3, "3D Viewer")
        self.setup_tab3()

        self.tab4 = QWidget()
        self.tabs.addTab(self.tab4, "Part Hanging")
        self.setup_tab4()

        # Connect new_file_loaded signal to Calculation tab
        self.step_analyzer.new_file_loaded.connect(self.Calculation.calc_powder_use)        
        self.step_analyzer.new_file_loaded.connect(self.Calculation.calculate_price)

        # Connect optimal and manual space to calculate price
        self.PartHanging.space_update.connect(self.Calculation.calculate_price)
        self.Calculation.calc_space_signal.connect(self.PartHanging.space)
        



    def setup_tab1(self):
        tab1_layout = QVBoxLayout()
        self.step_analyzer = STEPAnalyzer(self)
        tab1_layout.addWidget(self.step_analyzer)
        self.tab1.setLayout(tab1_layout)

        # Connect the signal to update the viewer in Tab 3
        self.step_analyzer.file_selected.connect(self.update_3d_viewer)


    def setup_tab2(self):
        tab2_layout = QVBoxLayout()
        colour_file_path = colourprice
        self.Calculation = Calculation(colour_file_path, parameter_path, self)
        tab2_layout.addWidget(self.Calculation)
        self.tab2.setLayout(tab2_layout)


    def setup_tab3(self):
        # Layout and content for the 3D Viewer
        tab3_layout = QVBoxLayout()
        self.viewer_widget = VTKViewer()  # Initialize without file initially
        tab3_layout.addWidget(self.viewer_widget)
        self.tab3.setLayout(tab3_layout)


    def setup_tab4(self):
        tab4_layout = QVBoxLayout()
        self.PartHanging = PartHanging(self)
        tab4_layout.addWidget(self.PartHanging)
        self.tab4.setLayout(tab4_layout)

    
    def update_3d_viewer(self, ply_file_path):
        # Load the new file in the 3D viewer widget
        self.viewer_widget.load_ply_file(ply_file_path)

    def get_bounding_box_dimensions(self):
        """
        Retrieve bounding box dimensions from STEPAnalyzer in Tab 1.
        Returns the dimensions as a list [XLength, YLength, ZLength].
        """
        # Access dimensions stored in the step_analyzer instance
        return self.step_analyzer.dimensions if self.step_analyzer else [0.0, 0.0, 0.0]
    

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec_())